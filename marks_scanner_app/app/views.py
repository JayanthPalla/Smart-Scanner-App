from django.shortcuts import render
from django.http import JsonResponse
import marks_scanner_app.settings as settings
from django.contrib.auth import authenticate, login
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from project_ML.main import ml_main
import os
from .models import CustomUser
from .serializers import CustomUserSerializer, UserRegistrationSerializer

# for excel generation
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Create your views here.

def index(request):
    return render(request, 'index.html')

##########################              USER AUTHENTICATION APIS                ##########################

@api_view(["POST"])
def register_user(request):
    print('form data:', request.data)
    if request.method == "POST":
        data = request.data
        serializer = UserRegistrationSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_409_CONFLICT)
    

@api_view(["POST"])
def login_user(request):
    if request.method == "POST":
        email = request.data.get('email')
        pswd = request.data.get('password')
        
        user = authenticate(request, username=email, password=pswd)
        
        print("User:", user)
        if user:
            login(request, user)
            return Response({'message':"Login Successful"}, status=status.HTTP_200_OK)
        return Response({'message':"Login Failed"}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_data(request):
    user = request.user
    user_info = {
        'id': user.id,
        'email': user.email,
        'full_name': user.fullName,
        'phone_number': user.phoneNumber,
        'dept': user.department,
    } 
    return Response(user_info, status=status.HTTP_200_OK)

@api_view(["GET"])
def check_login_status(request):
    is_logged_in = request.user.is_authenticated
    return Response({'is_logged_in': is_logged_in}, status=status.HTTP_200_OK)
  
   

################################    EXCEL GENERATION    ############################################

excel_text_alignment = Alignment(horizontal='center', vertical='center')

@api_view(['POST'])
def create_template(request):
    #creating a new workbook using workbook classs.....
    wb = Workbook()

    img = Image('app/static/images/rgukt.png')
    img.width = 73 
    img.height = 91 
    
    #creating the excel sheet
    ws = wb.active

    #adding the rgukt image at the 1st cell of the excel sheet (A1).
    ws.add_image(img, 'A1')

    #adding the title for the excel sheet which is common for all files.
    common_title = "Rajiv Gandhi University of knowledge Technologies"
    
    #font styles of the title rgukt:
    title_font = Font(name='Cambria', size=14, bold=True, italic=False, color='9c0507')

    #merging the required no.of cells for the rgukt text as same row we are merging startrow and endrow are same values.
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=12)

    #for single cell with dimensions along with title
    title_cell = ws.cell(row=1, column=2, value=common_title)

    #common style of font to be applied to  every cell
    title_cell.font = title_font

    #center-align the title within the merged cells
    title_cell.alignment = excel_text_alignment
    
    #row height setting....from the title---to----last of the title rgukt

    for row in range(1, 5):
        ws.row_dimensions[row].height = 20
    
    #-----------------------------------#
    
    ### Input parameters need to be taken from fronted React###
    
    Acad_year='2024-25'
    sem = request.data.get('semester')
    mid= request.data.get('mid')
    marks_sheet_title="A.Y.%s %s %s MARKS AWARD SHEETS"%(Acad_year, sem, mid)
    subject= request.data.get('subject')
    year = request.data.get('year')
    faculty_name = request.user.fullName
    branch = request.data.get('branch')
    
    
    # AY_2024-25_E3S1_CSE_OS_MID-1_marks_sheet
    template_filename = f'AY_{Acad_year}_{year}{sem}_{branch}_{subject}_{mid}_marks_sheet.xlsx'
    
    
    #----------------------------------#
    
    subtitles=["(Andhra Pradesh Government Act 18 of 2008)","Nuzvid, Andhra Pradesh – 521 202",marks_sheet_title]
    
    addr_font=Font(name='Cambria', size=14, bold=True, italic=False, color='000000')
    
    for i in range(len(subtitles)):
        ws.merge_cells(start_row=i+2, start_column=2, end_row=i+2, end_column=12)
        addr_title_cell = ws.cell(row=i+2, column=2, value=subtitles[i])
        addr_title_cell.font = addr_font
        addr_title_cell.alignment = excel_text_alignment
    
    
    
    #------------------------------------------#
    
    ### Input parameters need to be taken from fronted React###
    
    speacial=Font(name='Cambria', size=12, bold=True, italic=False, color='000000')
    
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
    addr_title_cell = ws.cell(row=6, column=2, value="SUBJECT NAME:%s"%(subject))
    addr_title_cell.font =  speacial
    addr_title_cell.alignment = excel_text_alignment
    
    ws.merge_cells(start_row=6, start_column=8, end_row=6, end_column=12)
    addr_title_cell = ws.cell(row=6, column=8, value="FACULTY NAME:%s"%(faculty_name))
    addr_title_cell.font = speacial
    addr_title_cell.alignment = excel_text_alignment
    
    #-------------------------------------#
    
    
    #Header values
    headers = {"S.NO":6, "ID":12, "Subject":0, "Subject Code":15, "Year":10, "Branch":10, "Class":10, "Class Code":13, "Acad Year":10, "Semester":10,"Marks":6}
    headers["Subject"]=len(subject)
    
    #font properties are:
    header_font = Font(name='Calibri', size=11, bold=True)
    col=1 #initial value of column
    for key,val in headers.items():
        cell=ws.cell(row=8,column=col,value=key)
        cell.font=header_font
        cell.alignment = excel_text_alignment
        #converting the column index to a letter
        col_letter = get_column_letter(col)
        #setting the column width based on the length of the value of the subject in the dictionary above!!!
        
        ws.column_dimensions[col_letter].width = val
        
        col+=1

    #----------------------------------------------------#
    
    # Save the template
    file_path = os.path.join(settings.MEDIA_ROOT, template_filename)
    wb.save(file_path)
    print(f"Template '{template_filename}' created successfully.")
    
    return JsonResponse({'response':f"Template '{template_filename}' created successfully."})
    

##############################      EXCEL FILE LOADING      ##########################################

#data font styles:
data_font= Font(name='Calibri', size=11, bold=False)
#--------------------------#

def add_student_data(ws, data_list):
    row_num = ws.max_row + 1  #here we get the maximum available empty row number
    for col_num, data in enumerate(data_list, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=data)
        cell.font = data_font
        cell.alignment = excel_text_alignment

#--------------------------#

student_data = []

@api_view(['POST'])
def load_excel_data(request):
    # Acad_year = '2024-25'
    # sem = 'semister-1' #inputs
    # mid = 'MID-2'
    # template_filename = f'AY_{Acad_year}_E3S{sem[-1]}_CSE3_{mid}_marks_sheet.xlsx'
    # subject = "OperatingSystems"
    subject_code = '20CS2309'
    # year = 'E3'
    # branch = 'CSE'
    classRoom = ['CSE-01', 'SG-06']
    
    Acad_year='2024-25'
    sem = request.data.get('semester')
    mid= request.data.get('mid')
    marks_sheet_title="A.Y.%s %s %s MARKS AWARD SHEETS"%(Acad_year, sem, mid)
    subject= request.data.get('subject')
    year = request.data.get('year')
    faculty_name = request.user.fullName
    branch = request.data.get('branch')
    
    
    # AY_2024-25_E3S1_CSE_OS_MID-1_marks_sheet
    template_filename = f'AY_{Acad_year}_{year}{sem}_{branch}_{subject}_{mid}_marks_sheet.xlsx'

    #Loading the xlsx file created from the previous program block.
    excel_file_path = os.path.join(settings.MEDIA_ROOT, template_filename)
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    if ws.max_row > 9:
        for i in ws.iter_rows(9, ws.max_row, values_only=True):
            student_data.append(list(i))
    
    j = 0
    while j < 60:
        # input_str = input("Enter ID and Marks (or '0' to exit): ")

        # if input_str == '0':
        #     break

        try:
            #taking the input in the ID,marks  format
            # student_id, marks = input_str.split(',')
            student_id, marks = ml_main()
            data_list = [0, student_id, subject, subject_code, year, branch, classRoom[0], classRoom[1], Acad_year, sem, int(marks)]
            print(data_list)
            student_data.append(data_list)
            j += 1
        except ValueError:
            print("Invalid input format. Please enter in 'ID,Marks' format.")
            continue

        add_student_data(ws, data_list)
        
        # display_excel_at_scan(ws)
        
    sort_rows_by_id(ws)

    #saving the changes to the Excel file:
    wb.save(excel_file_path)
    return JsonResponse({'response':'data loaded'})
    
    
def sort_rows_by_id(ws):
    ws.delete_rows(9, ws.max_row) #delete all existing rows from 9th row
    sorted_student_data = sorted(student_data, key=lambda x: x[1])
    
    roll=1
    for data_list in sorted_student_data:
        data_list[0]=roll
        add_student_data(ws, data_list)
        roll+=1
        
        
##############################      DISPLAY EXCEL IN REACT      ##########################################

# def display_excel_at_scan():
    